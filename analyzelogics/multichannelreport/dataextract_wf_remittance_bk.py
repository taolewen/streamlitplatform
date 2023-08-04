
#cd
import datetime
import os
import traceback
import uuid
from datetime import timedelta
from urllib.parse import quote_plus
import openpyxl
import pymysql
from dateutil import parser
import pandas as pd

# 显示所有列
from openpyxl import load_workbook
from sqlalchemy import create_engine


connstr = "mysql+pymysql://developer:%s@124.71.174.53:3306/csbd?charset=utf8" % quote_plus('csbd@123')
engine = create_engine(connstr)
pd.set_option('display.max_columns', None)
reporttype='dataextract_wf_remittance'

def getuid():
    uid = str(uuid.uuid4())
    suid = ''.join(uid.split('-'))
    return suid
def updatebatch(attrjson,batchid,path):
    conn = pymysql.connect(host='124.71.174.53',
                           user='developer',
                           password='csbd@123',
                           database='csbd')
    cursor = conn.cursor()
    sql = f"""insert newchannel_batchinfo (batchid,reporttype,path,area,country,week,store,qijian) values 
    ('{batchid}','{reporttype}','{path}','{attrjson['area'].upper()}','{(attrjson['country']).upper()}',
    '{attrjson['week']}','{attrjson['store'].upper()}','{attrjson['qijian']}')"""
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
def selectbatch():
    sql = f"""select * from newchannel_batchinfo where reporttype='{reporttype}' order by createdate desc"""
    df=pd.read_sql(sql,con=connstr)
    dl=df.to_dict('records')
    strlist=[]
    for d in dl:
        try:
            filename=d['path'].split('/')[-1]
        except:
            filename='notfound'
        str=f"{d['createdate']}_{filename}_{d['area']}_{d['country']}_{d['store']}_{d['week']}_{d['batchid']}"
        # print(str)
        strlist.append(str)
    return strlist
def deletebatch(batchid):
    try:
        conn = pymysql.connect(host='124.71.174.53',
                               user='developer',
                               password='csbd@123',
                               database='csbd')
        cursor = conn.cursor()
        sql0 = f"""delete from  newchannel_wf_credits where batchid = '{batchid}' """
        cursor.execute(sql0)
        sql1 = f"""delete from  newchannel_wf_earlypay where batchid = '{batchid}' """
        cursor.execute(sql1)
        sql2 = f"""delete from  newchannel_batchinfo where batchid = '{batchid}' """
        cursor.execute(sql2)
        try:
            sql3 = f"""delete from  newchannel_wf_earlypay_vendor_eu where batchid = '{batchid}' """
            cursor.execute(sql3)
        except:
            traceback.print_exc()
        sql4 = f"""delete from  newchannel_wf_earlypay_detail_us where batchid = '{batchid}' """
        cursor.execute(sql4)
        sql5 = f"""delete from  newchannel_wf_earlypay_detail_eu where batchid = '{batchid}' """
        cursor.execute(sql5)
        conn.commit()
        cursor.close()
        conn.close()
        return 1,''
    except:
        return 2,traceback.format_exc()
def dealsinglefile(path,attrjson):
    dfvendor=None
    try:
        batchid=getuid()

        excel=load_workbook(path)
        sheetnames=excel.get_sheet_names()
        print(sheetnames)
        table=excel.get_sheet_by_name(sheetnames[0])
        tablemaxrow=table.max_row

        blocklist={}
        currentl=1
        for i,cell in enumerate(table['A']):
            if not blocklist.get(currentl):
                blocklist[currentl]=[]
            # print(i+1,cell.value)
            if cell.value is None:
                currentl=currentl+1
                continue
            blocklist[currentl].append(i+1)
        print(blocklist)

        blocklist1={}#北美
        currentl=1
        for i,cell in enumerate(table['C']):
            if not blocklist1.get(currentl):
                blocklist1[currentl]=[]
            # print(i+1,cell.value)
            if cell.value is None:
                currentl=currentl+1
                continue
            blocklist1[currentl].append(i+1)
        print(blocklist1)

        blocklist2={} #欧洲
        currentl=1
        for i,cell in enumerate(table['I']):
            if not blocklist2.get(currentl):
                blocklist2[currentl]=[]
            # print(i+1,cell.value)
            if cell.value is None:
                currentl=currentl+1
                continue
            blocklist2[currentl].append(i+1)
        print(blocklist2)


        if attrjson['area'].upper() =='US':
            creditsdf=pd.DataFrame(columns=['po','credate','sku','qty','amount'])
            damageallowancerate, damageallowanceamount, earlypaydiscountrate, earlypaydiscountamount ,\
            jossmainallowancefordamagesrate,jossmainallowancefordamagesamount,jossmainearlypaydiscountrate,jossmainearlypaydiscountamount,\
            Castlegate_Fulfillment_Services,total= None, 0, None, 0,None, 0, None, 0, 0,0

            for key in blocklist.keys():
                if len(blocklist[key])!=0 and key >=3:

                    # print(table.cell(row=blocklist.get(key)[0],column=1).value)
                    if table.cell(row=blocklist.get(key)[0],column=1).value=='Wayfair Allowance for Damages/ Defects':
                        print('Wayfair Allowance for Damages/ Defects')
                        for i in blocklist.get(key):
                            if table.cell(row=i, column=1).value.startswith('Wayfair Allowance for Damages'):
                                damageallowancerate=table.cell(row=i, column=3).value
                                damageallowanceamount=table.cell(row=i, column=4).value

                            if table.cell(row=i, column=1).value.startswith('Wayfair Early Pay Discount'):
                                earlypaydiscountrate=table.cell(row=i, column=3).value
                                earlypaydiscountamount=table.cell(row=i, column=4).value

                            if table.cell(row=i, column=1).value.startswith('Joss & Main Allowance for Damages/ Defects'):
                                jossmainallowancefordamagesrate=table.cell(row=i, column=3).value
                                jossmainallowancefordamagesamount=table.cell(row=i, column=4).value

                            if table.cell(row=i, column=1).value.startswith('Joss & Main Early Pay Discount'):
                                jossmainearlypaydiscountrate=table.cell(row=i, column=3).value
                                jossmainearlypaydiscountamount=table.cell(row=i, column=4).value


                    if table.cell(row=blocklist.get(key)[0],column=1).value=='Vendor Services:':
                        print('Vendor Services')
                        Castlegate_Fulfillment_Services=0
                        for i in blocklist.get(key):
                            if table.cell(row=i, column=2).value is not None:
                                if table.cell(row=i, column=2).value.startswith('Castlegate Fulfillment Services'):
                                    Castlegate_Fulfillment_Services=Castlegate_Fulfillment_Services+table.cell(row=i, column=4).value

                    if table.cell(row=blocklist.get(key)[0], column=1).value == 'Credit' or table.cell(row=blocklist.get(key)[0], column=1).value == 'Deduction':
                        print('Credit')
                        po=table.cell(row=blocklist.get(key)[0], column=2).value
                        date=table.cell(row=blocklist.get(key)[0], column=3).value
                        amount=table.cell(row=blocklist.get(key)[0], column=4).value
                        # print(po,date,amount)
                        sku,qty=None,None

                        for i in blocklist.get(key):
                            if table.cell(row=i, column=1).value.startswith('Item'):
                                sku=table.cell(row=i, column=1).value.split(':')[1].replace(' ','')
                                qty=table.cell(row=i, column=2).value.split(':')[1].replace(' ','')

                        creditsdf.loc[len(creditsdf.index)] = [po,date,sku,qty,amount]
            # print(creditsdf)
            # creditsttl0=creditsdf.loc[creditsdf['amount']<0,:]['amount'].sum()
            # print('退货>>'+str(creditsttl0))

            for key in blocklist1.keys():
                if len(blocklist1[key]) != 0 and key >= 3:
                    if table.cell(row=blocklist1.get(key)[0], column=3).value == 'Total:':
                        print('total')
                        total=table.cell(row=blocklist1.get(key)[0], column=4).value



            # print(damageallowanceamount, damageallowancerate, earlypaydiscountamount, earlypaydiscountrate,total)

            rown=blocklist[2]
            # print(rown)
            names=['invoice','po','date','amount','storeid','ordertype']
            df=pd.read_excel(path,names=names,skiprows=rown[0]-1,skipfooter=tablemaxrow-rown[-1])
            # df.to_csv('remittance_test.csv')
            creditsdf2=df.loc[df['amount']<0,:][['po','date','amount']]
            creditsdf2['sku']='None'
            creditsdf2['qty']=1
            creditsdf2.rename(columns={'date':'credate'},inplace=True)
            creditsdf2=creditsdf2[['po','credate','sku','qty','amount']]
            creditsdfconcat=pd.concat([creditsdf,creditsdf2])
            creditsdfconcat['taxvat']=0
            creditsdfconcat.rename(columns={'amount':'credits'},inplace=True)

        else:

            rown=blocklist[4]
            print(rown)
            names=['invoice','po','invoice_date','product_amount','wfallowancefordamages','wfearlypaydiscount','shipping','other','taxvat','paymentamount','busniess','ordertype']
            df=pd.read_excel(path,names=names,skiprows=rown[0]-1,skipfooter=tablemaxrow-rown[-1])
            # df.to_csv('remittance_test.csv')
            taxvat=df['taxvat'].sum()
            creditshipping=df.loc[df['shipping']<0,:]['shipping'].sum()
            creditproductamount=df.loc[df['product_amount']<0,:]['product_amount'].sum()

            print('tax>>'+str(taxvat))
            print('creditshipping>>'+str(creditshipping))
            print('creditproductamount>>'+str(creditproductamount))

            creditsttl=creditshipping+creditproductamount
            creditsdfconcat=df[['po','invoice_date','product_amount','shipping','taxvat']]
            def ifgreat0then0(x):
                if x>0:
                    return 0
                else:
                    return x

            creditsdfconcat['product_amount']=creditsdfconcat['product_amount'].apply(lambda x:ifgreat0then0(x))
            creditsdfconcat['shipping']=creditsdfconcat['shipping'].apply(lambda x:ifgreat0then0(x))
            creditsdfconcat['credits']=creditsdfconcat.apply(lambda x: x.product_amount+x.shipping,axis=1)
            creditsdfconcat['qty']=creditsdfconcat.apply(lambda x:1 if x.credits !=0 else 0,axis=1)
            creditsdfconcat['sku']='None'
            creditsdfconcat.rename(columns={'invoice_date':'credate'},inplace=True)
            creditsdfconcat=creditsdfconcat[['po','credate','sku','qty','credits','taxvat']]
            earlypaydiscountamount=df['wfearlypaydiscount'].sum()
            damageallowanceamount=df['wfallowancefordamages'].sum()
            jossmainallowancefordamagesrate = 0
            jossmainallowancefordamagesamount = 0
            jossmainearlypaydiscountrate = 0
            jossmainearlypaydiscountamount = 0
            Castlegate_Fulfillment_Services=0



            for key in blocklist.keys():
                if len(blocklist[key])!=0 and key >=3:

                    if table.cell(row=blocklist.get(key)[0],column=1).value=='Vendor Services:':
                        rown1 = blocklist[key]
                        print(rown1)
                        dfvendor = pd.read_excel(path, skiprows=rown1[1] - 1, skipfooter=tablemaxrow - rown1[-1])
                        print(dfvendor)
                        dfvendor.rename(columns={'Invoice #': 'invoice', 'Program Type': 'program_type',
                                                 'Invoice Date': 'invoice_date', 'Description': 'desc',
                                                 'Amount': 'amount'}, inplace=True)
                        dfvendor = dfvendor[['invoice', 'program_type', 'invoice_date', 'desc', 'amount']]

                    if table.cell(row=blocklist.get(key)[0], column=1).value == 'Credit' or table.cell(row=blocklist.get(key)[0], column=1).value =='Deduction':
                        print('Credit')
                        po=table.cell(row=blocklist.get(key)[0], column=2).value
                        date=table.cell(row=blocklist.get(key)[0], column=3).value
                        amount=table.cell(row=blocklist.get(key)[0], column=10).value
                        # print(po,date,amount)
                        sku,qty=None,None

                        for i in blocklist.get(key):
                            if table.cell(row=i, column=1).value.startswith('Item'):
                                sku=table.cell(row=i, column=1).value.split(':')[1].replace(' ','')
                                qty=table.cell(row=i, column=2).value.split(':')[1].replace(' ','')

                        creditsdfconcat.loc[len(creditsdfconcat.index)] = [po,date,sku,qty,amount,0]

            #取total
            total=0
            for key in blocklist2.keys():
                if len(blocklist2[key])!=0:
                    if table.cell(row=blocklist2.get(key)[0], column=9).value == 'Total (GBP): ':
                        print('total')
                        total=table.cell(row=blocklist2.get(key)[0], column=10).value
            # if blocklist[7]!=[]:
            #     if table.cell(row=blocklist[7][0], column=1).value == 'Vendor Services:':
            #
            #         rown1=blocklist[7]
            #         print(rown1)
            #         dfvendor=pd.read_excel(path,skiprows=rown1[1]-1,skipfooter=tablemaxrow-rown1[-1])
            #         print(dfvendor)
            #         dfvendor.rename(columns={'Invoice #':'invoice','Program Type':'program_type','Invoice Date':'invoice_date','Description':'desc','Amount':'amount'},inplace=True)
            #         dfvendor=dfvendor[['invoice','program_type','invoice_date','desc','amount']]

        ##############
        if attrjson['area'].upper() == 'US':
            df['area'] = attrjson['area']
            df['country'] = attrjson['country']
            df['store'] = attrjson['store']
            df['week'] = attrjson['week']
            df['qijian'] = attrjson['qijian']
            df['batchid'] = batchid
            df.to_sql('newchannel_wf_earlypay_detail_us', con=engine, if_exists='append', index=False, index_label=False)

        else:
            df['area'] = attrjson['area']
            df['country'] = attrjson['country']
            df['store'] = attrjson['store']
            df['week'] = attrjson['week']
            df['qijian'] = attrjson['qijian']
            df['batchid'] = batchid
            df.to_sql('newchannel_wf_earlypay_detail_eu', con=engine, if_exists='append', index=False, index_label=False)
            if dfvendor is not None:
                dfvendor['area'] = attrjson['area']
                dfvendor['country'] = attrjson['country']
                dfvendor['store'] = attrjson['store']
                dfvendor['week'] = attrjson['week']
                dfvendor['qijian'] = attrjson['qijian']
                dfvendor['batchid'] = batchid
                dfvendor.to_sql('newchannel_wf_earlypay_vendor_eu', con=engine, if_exists='append', index=False, index_label=False)

        creditsdfconcat['area']=attrjson['area']
        creditsdfconcat['country']=attrjson['country']
        creditsdfconcat['store']=attrjson['store']
        creditsdfconcat['week']=attrjson['week']
        creditsdfconcat['qijian'] = attrjson['qijian']

        creditsdfconcat=creditsdfconcat[['area','country','store','week','qijian','po','credate','sku','qty','credits','taxvat']]
        creditsdfconcat['batchid']=batchid

        creditsdfconcat.to_sql('newchannel_wf_credits', con=engine, if_exists='append', index=False, index_label=False)

        dfearlypay=pd.DataFrame(columns=['area','country','store','week','qijian','earlypaydiscountamount','allowancefordamagesdefects','jossmainallowancefordamagesamount','jossmainearlypaydiscountamount','Castlegate_Fulfillment_Services','total'])
        dfearlypay.loc[len(dfearlypay.index)] = [attrjson['area'], attrjson['country'], attrjson['store'], attrjson['week'],attrjson['qijian'], earlypaydiscountamount,damageallowanceamount,jossmainallowancefordamagesamount,jossmainearlypaydiscountamount,Castlegate_Fulfillment_Services,total]
        dfearlypay['batchid']=batchid

        dfearlypay.to_sql('newchannel_wf_earlypay', con=engine, if_exists='append', index=False, index_label=False)
        updatebatch(attrjson,batchid,path)

        return 1,''
    except:
        return 2,traceback.format_exc()


if __name__ == '__main__':
    # attrjson={
    #     'area': 'us',
    #     'country':'',
    #     'store':'WF-US-1',
    #     'week':20,
    #     'importdate':'2022-03-06',
    #       'qijian': '2022-15'
    #
    # }
    # print(dealsinglefile(
    # 'D:\pythonws\pythonws\playwrighttest\data_proceed\csvs\wfremittance\\us-Wayfair_Remittance_4416132.xlsx',attrjson)
    #     # '/data_proceed/csvs/wfremittance/us-Wayfair_Remittance_4434241.xlsx',attrjson)
    # )

    attrjson={
        'area': 'eu',
        'country':'',
        'store':'WF-EU-5',
        'week':20,
        'importdate':'2022-03-20',
        'qijian':'2022-15'

    }
    print(

    dealsinglefile(
    'D:\pythonws\pythonws\playwrighttest\data_proceed\csvs\wfremittance\eu-wwayfair_remittance_20221025.xlsx',attrjson)

    )


