
#cd
import datetime
import os
import traceback
import uuid
from datetime import timedelta
from urllib.parse import quote_plus
import openpyxl
import pymysql
from dateutil import parser
import pandas as pd

# 显示所有列
from openpyxl import load_workbook
from sqlalchemy import create_engine

import streamlit as st
host = st.secrets["mysql"]['host'],
user = st.secrets["mysql"]['user'],
password = st.secrets["mysql"]['password'],
db = st.secrets["mysql"]['database']
connstr = f"mysql+pymysql://{user[0]}:%s@{host[0]}:3306/{db}?charset=utf8" % quote_plus(f'{password[0]}')

engine = create_engine(connstr)
pd.set_option('display.max_columns', None)
reporttype='dataextract_cd_orderextract'
def getuid():
    uid = str(uuid.uuid4())
    suid = ''.join(uid.split('-'))
    return suid
def updatebatch(attrjson,batchid,path):
    conn = pymysql.connect(host=st.secrets["mysql"]['host'],
                                user=st.secrets["mysql"]['user'],
                                password=st.secrets["mysql"]['password'],
                                db=st.secrets["mysql"]['database'])
    cursor = conn.cursor()
    sql = f"""insert newchannel_batchinfo (batchid,reporttype,path,area,country,week,store,qijian) values 
    ('{batchid}','{reporttype}','{path}','{attrjson['area'].upper()}','{(attrjson['country']).upper()}',
    '{attrjson['week']}','{attrjson['store'].upper()}','{attrjson['qijian']}')"""
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
def selectbatch(attrjson):
    sql = f"""select * from newchannel_batchinfo where reporttype='{reporttype}' 
    {f'''and area='{attrjson['area']}' ''' if attrjson['area'] else ''}
    {f'''and country='{attrjson['country']}' ''' if attrjson['country'] else ''}
    {f'''and store='{attrjson['store']}' ''' if attrjson['store'] else ''}
    {f'''and week='{attrjson['week']}' ''' if attrjson['week'] else ''}
    {f'''and qijian='{attrjson['qijian']}' ''' if attrjson['qijian'] else ''}
    order by createdate desc"""
    df=pd.read_sql(sql,con=connstr)
    dl=df.to_dict('records')
    strlist=[]
    for d in dl:
        try:
            filename=d['path'].split('/')[-1]
        except:
            filename='notfound'
        str=f"{d['createdate']}_{filename}_{d['area']}_{d['country']}_{d['store']}_{d['week']}_{d['batchid']}"
        # print(str)
        strlist.append(str)
    def getfilename(x):
        try:
            return x.split('/')[-1]
        except:
            return 'none'
    df['filename']=df.apply(lambda x:getfilename(x.path),axis=1)
    df=df.drop('path', axis=1)
    df['delete']=False
    df=df[['delete','area','country','qijian','week','store','filename','createdate','reporttype','batchid']]
    return df
def deletebatch(batchid):
    try:
        conn = pymysql.connect(host=st.secrets["mysql"]['host'],
                               user=st.secrets["mysql"]['user'],
                               password=st.secrets["mysql"]['password'],
                               db=st.secrets["mysql"]['database'])
        cursor = conn.cursor()
        sql0 = f"""delete from  newchannel_cd_orderextract where batchid = '{batchid}' """
        cursor.execute(sql0)
        sql = f"""delete from  newchannel_batchinfo where batchid = '{batchid}' """
        cursor.execute(sql)
        conn.commit()
        cursor.close()
        conn.close()
        return 1,''
    except:
        return 2,traceback.format_exc()
def dealsinglefile(path,attrjson):
    try:
        batchid=getuid()
        excel=load_workbook(path)
        sheetnames=excel.get_sheet_names()
        print(sheetnames)
        table=excel.get_sheet_by_name(sheetnames[0])
        shop=table.cell(row=1,column=2).value
        fromdate=table.cell(row=1,column=5).value
        todate=table.cell(row=1,column=7).value
        export_date=table.cell(row=4,column=2).value
        print(shop,fromdate,todate,export_date)
        name=['Corporation','order_id','order_date','last_modificationdate','order_status','shipping_mode','product_details','product_condition',
              'seller_reference','EAN','sku_cdiscount','product_status','quantity','unit_price','shipping_fee','commission',
              'seller_income','civility','last_name','first_name','address','shippingzipcode','shippingcity','shippingcountry','shippingphone1','shippingphone2','billingnameandaddress','billingzipcode','billingcity','billingcountry']
        df=pd.read_excel(path,names=name,skiprows=5)
        df['shop']=shop
        df['daterange']=fromdate+'~'+todate
        df['exportdate']=export_date
        df['qijian']=attrjson['qijian']
        df['country']=(attrjson['country']).upper()
        df['week']=attrjson['week']
        df['store']=attrjson['store'].upper()
        df['area']=attrjson['area'].upper()
        df['batchid']=batchid
        print(df)
        print(df.shape)
        df=df[['area','country','store','week','qijian',
            'Corporation','order_id','order_date','last_modificationdate','order_status','shipping_mode','product_details','product_condition',
              'seller_reference','EAN','sku_cdiscount','product_status','quantity','unit_price','shipping_fee','commission',
              'seller_income','civility','last_name','first_name','address','shippingzipcode','shippingcity','shippingcountry','shippingphone1','shippingphone2','billingnameandaddress','billingzipcode','billingcity','billingcountry','batchid']]
        # df.to_sql()

        df.to_sql(name='newchannel_cd_orderextract', con=engine, if_exists='append', index=False, index_label=False)
        updatebatch(attrjson,batchid,path)
        return 1,''
    except:
        return 2,traceback.format_exc()
if __name__ == '__main__':
    # attrjson={
    #     'country':'fr',
    #     'store':'cd-3'
    # }
    # dealsinglefile('E:\pythonws\pythonws\pythonws\playwrighttest\data_proceed\csvs\cdorderextract\示例-OrderExtract_KSBD_2021-03-06.xlsx',attrjson)
    selectbatch()




