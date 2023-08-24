
#cd
import datetime
import os
import traceback
import uuid
from datetime import timedelta
from urllib.parse import quote_plus
import openpyxl
import pymysql
from dateutil import parser
import pandas as pd

# 显示所有列
from openpyxl import load_workbook
from sqlalchemy import create_engine


connstr = "mysql+pymysql://developer:%s@124.71.174.53:3306/csbd?charset=utf8" % quote_plus('csbd@123')
engine = create_engine(connstr)
pd.set_option('display.max_columns', None)
reporttype='dataextract_cd_paymentdetail_returnmoney'

def getuid():
    uid = str(uuid.uuid4())
    suid = ''.join(uid.split('-'))
    return suid
def updatebatch(attrjson,batchid,path):
    conn = pymysql.connect(host='124.71.174.53',
                           user='developer',
                           password='csbd@123',
                           database='csbd')
    cursor = conn.cursor()
    sql = f"""insert newchannel_batchinfo (batchid,reporttype,path,area,country,week,store,qijian) values 
    ('{batchid}','{reporttype}','{path}','{attrjson['area'].upper()}','{(attrjson['country']).upper()}',
    '{attrjson['week']}','{attrjson['store'].upper()}','{attrjson['qijian']}')"""
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
def selectbatch(attrjson):
    sql = f"""select * from newchannel_batchinfo where reporttype='{reporttype}' 
    {f'''and area='{attrjson['area']}' ''' if attrjson['area'] else ''}
    {f'''and country='{attrjson['country']}' ''' if attrjson['country'] else ''}
    {f'''and store='{attrjson['store']}' ''' if attrjson['store'] else ''}
    {f'''and week='{attrjson['week']}' ''' if attrjson['week'] else ''}
    {f'''and qijian='{attrjson['qijian']}' ''' if attrjson['qijian'] else ''}
    order by createdate desc"""
    df=pd.read_sql(sql,con=connstr)
    dl=df.to_dict('records')
    strlist=[]
    for d in dl:
        try:
            filename=d['path'].split('/')[-1]
        except:
            filename='notfound'
        str=f"{d['createdate']}_{filename}_{d['area']}_{d['country']}_{d['store']}_{d['week']}_{d['batchid']}"
        # print(str)
        strlist.append(str)
    df = df.drop('path', axis=1)
    df['delete']=False
    df=df[['delete','area','country','qijian','week','store','createdate','reporttype','batchid']]
    return df
def deletebatch(batchid):
    try:
        conn = pymysql.connect(host='124.71.174.53',
                               user='developer',
                               password='csbd@123',
                               database='csbd')
        cursor = conn.cursor()
        sql0 = f"""delete from  newchannel_cd_paymentdetail where batchid = '{batchid}' """
        cursor.execute(sql0)
        sql = f"""delete from  newchannel_batchinfo where batchid = '{batchid}' """
        cursor.execute(sql)
        conn.commit()
        cursor.close()
        conn.close()
        return 1,''
    except:
        return 2,traceback.format_exc()
def dealsinglefile(path,attrjson):
    try:
        batchid=getuid()
        #添加一行否则第一行导不进去
        excel = load_workbook(path)
        sheetnames = excel.sheetnames
        table = excel[sheetnames[0]]
        table.insert_rows(5)
        excel.save(path)
        #######
        name=['Canal_de_vente','ID_Boutique','N_facture_avoir','N_commande_Service','Date_opération_comptable','Date_de_mise_en_paiement_anticipé',
              'Date_virement','Vente_TTC_hors_frais_de_port','Frais_de_port_TTC','Commission_Produit','Commission_Facilités_de_paiement','Commission_Frais_de_paiement_4_fois',
              'Remboursement_TTC_hors_frais_de_port','Avoir_commission','Montant_de_TVA_sur_Vente_remboursement','Montant_de_TVA_sur_frais_de_port','Total_recu','Devise']
        #第一行导不进去
        df=pd.read_excel(path,names=name,skiprows=3,thousands=',')
        df['qijian']=attrjson['qijian']
        df['country']=(attrjson['country']).upper()
        df['week']=attrjson['week']
        df['store']=attrjson['store'].upper()
        df['area']=attrjson['area'].upper()
        df['batchid']=batchid

        print(df)
        print(df.shape)
        # df.to_sql()
        df=df[[
            'area', 'country', 'store', 'week', 'qijian',
            'Canal_de_vente','ID_Boutique','N_facture_avoir','N_commande_Service','Date_opération_comptable','Date_de_mise_en_paiement_anticipé',
              'Date_virement','Vente_TTC_hors_frais_de_port','Frais_de_port_TTC','Commission_Produit','Commission_Facilités_de_paiement','Commission_Frais_de_paiement_4_fois',
              'Remboursement_TTC_hors_frais_de_port','Avoir_commission','Montant_de_TVA_sur_Vente_remboursement','Montant_de_TVA_sur_frais_de_port','Total_recu','Devise','batchid']]

        df.to_sql(name='newchannel_cd_paymentdetail', con=engine, if_exists='append', index=False, index_label=False)
        updatebatch(attrjson,batchid,path)

        return 1,''
    except:
        return 2,traceback.format_exc()
if __name__ == '__main__':
    attrjson={
        'country':'fr',
        'store':'cd-3'
    }
    dealsinglefile('E:\pythonws\pythonws\pythonws\playwrighttest\data_proceed\csvs\cdpaymentdetail\KSBD-3-payment_details_export_107239.xlsx',attrjson)
    # dealsinglefile('E:\pythonws\pythonws\pythonws\playwrighttest\data_proceed\csvs\cdpaymentdetail\\NSD-payment_details_export_139494.xlsx',attrjson)





